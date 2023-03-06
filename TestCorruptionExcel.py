import pandas as pd
import openpyxl
from openpyxl import load_workbook


#Chargement fichier Excel
filename = 'test/bddTest.xlsx'
filename2 = 'test/bddTest - Copie.xlsx'
wb = load_workbook(filename)

#Extraction onglet requete
for sheet_name in wb.sheetnames:
    print(sheet_name)
    if sheet_name != 'Req':
        sheet = wb[sheet_name]
        colonnesees = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        print(colonnesees)
        
        colonne_prime = None
        print(list(sheet.values))
        iColonne = 0
        for colonne in sheet.iter_cols(min_col=1, values_only=True):
            iColonne +=1
            if colonne[0] == "Prime":
                colonne_prime = list(colonne)
                break

        if colonne_prime is not None:
            for i in range(1, len(colonne_prime)):
                colonne_prime[i] += 1
            
            for i, valeur in enumerate(colonne_prime):
                sheet.cell(row=i+1, column=iColonne, value=valeur)
        print("nouvelle list")
        print(list(sheet.values))
#Chargement des autres onglets dans les pandas

##TEST PANDAS###

xlsx_file = pd.ExcelFile(filename2)

Req_sheet = pd.read_excel(xlsx_file, sheet_name='Req')
df1 = pd.read_excel(xlsx_file, sheet_name='NomPrenom')
df2 = pd.read_excel(xlsx_file, sheet_name='Revenus')

#modif
print(df2)
df2['Prime'] = df2['Prime'].apply(lambda x: x+1)
print("modif")
print(df2)

#with pd.ExcelWriter(filename2, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
with pd.ExcelWriter(filename2) as writer:
    #Req_sheet.to_excel(writer, sheet_name='Req', index=False)

    df1.to_excel(writer, sheet_name='NomPrenom', index=False, header=True)
    df2.to_excel(writer, sheet_name='Revenus', index=False, header=True)
#Ecriture dans les onglets
wb.save(filename)